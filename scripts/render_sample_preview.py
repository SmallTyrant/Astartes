"""Render PDF뷰어_ios tab of goodnotes.xlsx as a PNG preview.

Output: docs/images/sample-preview.png
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "outputs" / "sheets" / "goodnotes.xlsx"
OUT = ROOT / "docs" / "images" / "sample-preview.png"

FONT_REG = "/System/Library/Fonts/AppleSDGothicNeo.ttc"
FONT_BOLD = "/System/Library/Fonts/AppleSDGothicNeo.ttc"

SCALE = 4
PAD = 8 * SCALE
LINE_H = 18 * SCALE
HEADER_H = 22 * SCALE
ROW_MIN_H = 60 * SCALE

COL_WIDTHS = [
    ("TC ID", 70),
    ("priority", 80),
    ("1 Step", 150),
    ("2 Step", 150),
    ("3 Step", 150),
    ("4 Step", 150),
    ("5 Step", 150),
    ("pre-condition", 130),
    ("기대결과", 200),
    ("result", 70),
    ("Jira ticket", 70),
]

HEADER_FILL = (181, 215, 168)
PRIO_HIGH = (224, 102, 102)
PRIO_MID = (255, 217, 102)
PRIO_LOW = (147, 196, 125)
BORDER = (191, 191, 191)
TEXT = (0, 0, 0)
WHITE = (255, 255, 255)


def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    ws = wb["PDF뷰어_ios"]
    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        rows.append(row[1:12])  # B..L → TC ID..Jira ticket
    return rows


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(FONT_BOLD if bold else FONT_REG, size)


def wrap_text(draw: ImageDraw.ImageDraw, text: str, max_w: int, fnt: ImageFont.FreeTypeFont) -> list[str]:
    if text is None or text == "":
        return [""]
    text = str(text)
    out: list[str] = []
    for paragraph in text.split("\n"):
        line = ""
        for ch in paragraph:
            test = line + ch
            w = draw.textlength(test, font=fnt)
            if w > max_w and line:
                out.append(line)
                line = ch
            else:
                line = test
        out.append(line)
    return out


def main() -> None:
    rows = load_rows()
    col_pix = [w * SCALE for _, w in COL_WIDTHS]
    table_w = sum(col_pix)
    margin = 20 * SCALE
    canvas_w = table_w + margin * 2

    fnt_header = font(11 * SCALE, bold=True)
    fnt_body = font(10 * SCALE)
    fnt_prio = font(11 * SCALE, bold=True)

    tmp = Image.new("RGBA", (10, 10), WHITE)
    tdraw = ImageDraw.Draw(tmp)

    row_heights = []
    for r in rows:
        max_lines = 1
        for c, val in enumerate(r):
            if c == 0 or c == 1:
                continue
            lines = wrap_text(tdraw, val if val is not None else "", col_pix[c] - PAD * 2, fnt_body)
            if len(lines) > max_lines:
                max_lines = len(lines)
        h = max(ROW_MIN_H, PAD * 2 + max_lines * LINE_H)
        row_heights.append(h)

    canvas_h = margin * 2 + HEADER_H + sum(row_heights)
    img = Image.new("RGBA", (canvas_w, canvas_h), WHITE)
    draw = ImageDraw.Draw(img)

    x0 = margin
    y0 = margin

    # Header row
    x = x0
    draw.rectangle([x0, y0, x0 + table_w, y0 + HEADER_H], fill=HEADER_FILL, outline=BORDER, width=SCALE)
    for i, (label, _) in enumerate(COL_WIDTHS):
        cx = x + col_pix[i] / 2
        cy = y0 + HEADER_H / 2
        tw = draw.textlength(label, font=fnt_header)
        bbox = fnt_header.getbbox(label)
        th = bbox[3] - bbox[1]
        draw.text((cx - tw / 2, cy - th / 2 - bbox[1]), label, font=fnt_header, fill=TEXT)
        if i > 0:
            draw.line([x, y0, x, y0 + HEADER_H], fill=BORDER, width=SCALE // 2 or 1)
        x += col_pix[i]

    # Body rows
    y = y0 + HEADER_H
    for r_idx, r in enumerate(rows):
        h = row_heights[r_idx]
        # row outline
        draw.rectangle([x0, y, x0 + table_w, y + h], outline=BORDER, width=SCALE // 2 or 1)
        # priority fill
        prio = (r[1] or "").lower() if r[1] else ""
        prio_color = {"high": PRIO_HIGH, "mid": PRIO_MID, "low": PRIO_LOW}.get(prio, WHITE)
        prio_x = x0 + col_pix[0]
        draw.rectangle([prio_x, y, prio_x + col_pix[1], y + h], fill=prio_color, outline=BORDER, width=SCALE // 2 or 1)

        cx = x0
        for c, val in enumerate(r):
            cell_w = col_pix[c]
            if c == 0:
                txt = "" if val is None else str(int(val)) if isinstance(val, (int, float)) else str(val)
                tw = draw.textlength(txt, font=fnt_body)
                bbox = fnt_body.getbbox(txt or "X")
                th = bbox[3] - bbox[1]
                draw.text((cx + cell_w / 2 - tw / 2, y + h / 2 - th / 2 - bbox[1]), txt, font=fnt_body, fill=TEXT)
            elif c == 1:
                txt = "" if val is None else str(val)
                tw = draw.textlength(txt, font=fnt_prio)
                bbox = fnt_prio.getbbox(txt or "X")
                th = bbox[3] - bbox[1]
                text_color = WHITE if prio in ("high", "low") else TEXT
                draw.text((cx + cell_w / 2 - tw / 2, y + h / 2 - th / 2 - bbox[1]), txt, font=fnt_prio, fill=text_color)
            else:
                lines = wrap_text(draw, val if val is not None else "", cell_w - PAD * 2, fnt_body)
                ty = y + PAD
                for line in lines:
                    draw.text((cx + PAD, ty), line, font=fnt_body, fill=TEXT)
                    ty += LINE_H

            # vertical sep
            if c > 0:
                draw.line([cx, y, cx, y + h], fill=BORDER, width=SCALE // 2 or 1)
            cx += cell_w
        y += h

    # downscale 50% for crisp anti-aliasing
    final = img.resize((canvas_w // 2, canvas_h // 2), Image.LANCZOS)
    final.save(OUT, format="PNG", optimize=True)
    print(f"wrote {OUT} {final.size}")


if __name__ == "__main__":
    main()
