#!/usr/bin/env python3
"""TC JSON(v3) → 단일 XLSX 워크북 (앱당 1파일).

레이아웃 (참조 시트 정합):
  - 시트 순서: summury, {screen}_{platform}, ...
  - TC 시트:
    - A 컬럼 비움
    - Row 1 비움
    - Row 2 헤더 (B~M): TC ID | priority | 1 Step~5 Step | pre-condition | 기대결과 | result | Jira ticket | 비고
    - Row 3+ 데이터
    - 헤더 배경: 연두 (#B6D7A8)
    - priority 컬러: high=빨강, mid=노랑, low=초록 (조건부 서식)
    - result: 드롭다운 (Pass/Fail/Block/N/A), 기본값 빈 칸 (QA가 실행 후 채움)
  - summury 시트:
    - B2: 앱명
    - B4:E7 환경 블록 (수행 환경/수행자/테스트 기간) — 빈 칸, QA가 기입
    - B9:J9 통계 헤더 (회색 #D9D9D9)
    - B10~ 각 (screen_platform) 시트별 COUNTA/COUNTIF 수식
    - 마지막 행: 총 합계

사용:
  python3 .claude/skills/astartes-tc/scripts/export_workbook.py [앱명] [--spec-version v1.2]
  - 앱명 미지정 시 'app' 사용
  - --spec-version: 명세서 버전 (지정 시 삭제 TC 비고에 함께 기록)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# 실행 시점의 작업 디렉토리를 프로젝트 루트로 사용.
# 전역 설치(~/.claude/skills/) 환경에서도 올바른 outputs/ 경로를 얻기 위해 cwd() 사용.
ROOT = Path.cwd()
SRC_DIR = ROOT / "outputs" / "testcases"
OUT_DIR = ROOT / "outputs" / "sheets"

ALLOWED_PLATFORM = {"and", "ios", "web"}

TC_HEADER = [
    "TC ID", "priority",
    "1 Step", "2 Step", "3 Step", "4 Step", "5 Step",
    "pre-condition", "기대결과", "result", "Jira ticket", "비고",
]
HEADER_START_COL = 2   # B
HEADER_ROW = 2
DATA_START_ROW = 3

# 헤더 순서 기반 컬럼 자동 계산 — 헤더 변경 시 자동 반영
PRIORITY_COL = get_column_letter(HEADER_START_COL + TC_HEADER.index("priority"))  # C
RESULT_COL   = get_column_letter(HEADER_START_COL + TC_HEADER.index("result"))    # K

GREEN = PatternFill("solid", fgColor="B6D7A8")
GRAY = PatternFill("solid", fgColor="D9D9D9")
RED = PatternFill("solid", fgColor="E06666")
YELLOW = PatternFill("solid", fgColor="FFD966")
LIGHT_GREEN = PatternFill("solid", fgColor="93C47D")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
HEADER_FONT = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")

CONTENT_FIELDS = ("title", "steps", "expected", "precondition", "priority", "risk_tags")
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?\[\]:]")


# ── 유틸리티 ──────────────────────────────────────────────────────────────────

def _content_hash(tc: dict) -> str:
    payload = {f: tc.get(f) for f in CONTENT_FIELDS}
    return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode()).hexdigest()


def normalize_sheet_name(name: str, existing: set[str]) -> str:
    """Excel 31자 제한 준수 + 금지 문자 제거 + 중복 회피."""
    name = _INVALID_SHEET_CHARS.sub("_", name)[:31]
    if name not in existing:
        return name
    for i in range(1, 100):
        candidate = f"{name[:28]}_{i}"
        if candidate not in existing:
            return candidate
    return name  # fallback (사실상 도달 불가)


# ── 스냅샷 I/O ────────────────────────────────────────────────────────────────

def extract_result_snapshot(xlsx_path: Path) -> dict[str, dict]:
    """사이드카 JSON에서 {tc_id: {result, content_hash, tc_data}} 스냅샷을 로드한다.

    신규 포맷: value가 dict (또는 빈 dict) → 그대로 반환.
    구 포맷:   value가 str (hash 문자열) → XLSX에서 result를 보완해 반환.
    """
    if not xlsx_path.exists():
        return {}

    sidecar = xlsx_path.with_name(xlsx_path.stem + "_snapshot.json")
    hash_map: dict[str, str] = {}

    if sidecar.exists():
        try:
            raw = json.loads(sidecar.read_text(encoding="utf-8"))
            # 신규 포맷: dict이고 (비어있거나 첫 value가 dict)
            if isinstance(raw, dict) and (not raw or isinstance(next(iter(raw.values())), dict)):
                return raw
            # 구 포맷: value가 hash 문자열
            if isinstance(raw, dict):
                hash_map = raw
        except Exception as e:
            print(f"[export_workbook] 사이드카 JSON 파싱 실패 → result 보존 불가: {e}", file=sys.stderr)

    # XLSX에서 result 읽기 (구 포맷 호환 또는 사이드카 없는 경우)
    result_map: dict[str, dict] = {}
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                if ws.title == "summury":
                    continue
                for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
                    tc_id_val = row[HEADER_START_COL - 1] if len(row) >= HEADER_START_COL else None
                    result_val = row[HEADER_START_COL - 1 + 9] if len(row) >= HEADER_START_COL + 9 else None
                    if tc_id_val is None:
                        continue
                    tc_id = str(tc_id_val)
                    result_map[tc_id] = {
                        "result": str(result_val) if result_val else "",
                        "content_hash": hash_map.get(tc_id, ""),
                        "tc_data": None,  # 구 포맷 — 삭제 TC 복원 불가
                    }
        finally:
            wb.close()
    except Exception as e:
        print(f"[export_workbook] 기존 XLSX 읽기 실패 (스냅샷 무시): {e}", file=sys.stderr)
    return result_map


def merge_results(
    tcs: list[dict],
    snapshot: dict[str, dict],
    deletion_note: str = "명세 변경으로 삭제됨",
) -> list[dict]:
    """TC 목록에 기존 result를 병합한다.

    - 내용 무변경: result 복원
    - 내용 변경:   result 초기화
    - 삭제된 TC:   N/A + 비고에 사유 기록 (tc_data가 있을 때만)
    """
    if not snapshot:
        return tcs

    # tc_id=None 인 TC는 집합에서 제외해 오배정 방지
    new_ids = {str(tc.get("tc_id")) for tc in tcs if tc.get("tc_id") is not None}
    merged = []

    for tc in tcs:
        tc_id_raw = tc.get("tc_id")
        if tc_id_raw is not None:
            tc_id = str(tc_id_raw)
            if tc_id in snapshot:
                snap = snapshot[tc_id]
                old_hash = snap.get("content_hash", "")
                new_hash = _content_hash(tc)
                if old_hash == new_hash:
                    tc = {**tc, "result": snap.get("result", "")}
                else:
                    tc = {**tc, "result": ""}
                    print(f"[export_workbook] {tc_id} 내용 변경 → result 초기화", file=sys.stderr)
        merged.append(tc)

    # 삭제된 TC → N/A 유지
    for tc_id, snap in snapshot.items():
        if tc_id not in new_ids:
            tc_data = snap.get("tc_data")
            if tc_data:
                merged.append({**tc_data, "result": "N/A", "note": deletion_note})
                print(f"[export_workbook] {tc_id} 명세에서 삭제됨 → N/A 유지", file=sys.stderr)

    return merged


def save_snapshot(xlsx_path: Path, all_tcs: list[dict]) -> None:
    """tc_id → {content_hash, result, tc_data} 사이드카 JSON을 원자적으로 저장한다."""
    sidecar = xlsx_path.with_name(xlsx_path.stem + "_snapshot.json")
    tmp = sidecar.with_suffix(".tmp")
    data = {
        str(tc.get("tc_id", "")): {
            "content_hash": _content_hash(tc),
            "result": tc.get("result", ""),
            "tc_data": tc,
        }
        for tc in all_tcs
    }
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(sidecar)  # atomic rename — XLSX 저장 실패와 스냅샷 불일치 방지


# ── TC 로딩 / 그룹화 ──────────────────────────────────────────────────────────

def load_json(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and "testcases" in data:
        data = data["testcases"]
    if not isinstance(data, list):
        data = [data]
    return [tc for tc in data if isinstance(tc, dict)]


def group_by_tab(tcs: list[dict]) -> dict[tuple[str, str], list[dict]]:
    groups: dict[tuple[str, str], list[dict]] = {}
    for tc in tcs:
        screen = (tc.get("screen") or "").strip()
        platform = (tc.get("platform") or "").strip()
        if not screen or platform not in ALLOWED_PLATFORM:
            print(f"[export_workbook] skip: screen={screen!r} platform={platform!r}", file=sys.stderr)
            continue
        groups.setdefault((screen, platform), []).append(tc)
    for lst in groups.values():
        # tc_id 없는 시드/메모 행은 맨 뒤로 보낸다 (정렬 안정성 + str 캐스트로 int/str 혼재 시 TypeError 방지)
        lst.sort(key=lambda t: (t.get("tc_id") in (None, ""), str(t.get("tc_id") or "")))
    return groups


def steps_to_cells(steps: list) -> list[str]:
    out = list(steps or [])
    while len(out) < 5:
        out.append("")
    return out[:5]


# ── 시트 생성 ────────────────────────────────────────────────────────────────

def add_tc_sheet(wb: Workbook, sheet_name: str, tcs: list[dict]) -> None:
    """sheet_name은 normalize_sheet_name()으로 이미 정규화된 값이어야 한다."""
    ws = wb.create_sheet(title=sheet_name)

    widths = {1: 3, 2: 7, 3: 11, 4: 28, 5: 28, 6: 28, 7: 28, 8: 28, 9: 24, 10: 32, 11: 9, 12: 14, 13: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, label in enumerate(TC_HEADER):
        cell = ws.cell(row=HEADER_ROW, column=HEADER_START_COL + i, value=label)
        cell.fill = GREEN
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    prev_steps: list[str] | None = None
    for idx, tc in enumerate(tcs):
        row = DATA_START_ROW + idx
        steps = steps_to_cells(tc.get("steps") or [])
        # 시각적 공백: 직전 행의 같은 컬럼과 동일하면 빈 칸으로 둔다 (PDF뷰어_and 패턴).
        # 상태 변경 단계만 노출되어 시트가 step-progression처럼 읽힌다.
        display_steps = list(steps)
        if prev_steps is not None:
            for j in range(5):
                if display_steps[j] and display_steps[j] == prev_steps[j]:
                    display_steps[j] = ""
        prev_steps = steps
        # tc_id가 None이면 빈 칸 유지 (시드/메모 행). priority도 동일.
        tc_id_val = tc.get("tc_id")
        values = [
            tc_id_val if tc_id_val not in (None, "") else None,
            tc.get("priority") or None,
            display_steps[0], display_steps[1], display_steps[2], display_steps[3], display_steps[4],
            tc.get("precondition", "") or "",
            tc.get("expected", "") or "",
            tc.get("result", "") or "",
            tc.get("jira_ticket", "") or "",
            tc.get("note", "") or "",
        ]
        for j, val in enumerate(values):
            cell = ws.cell(row=row, column=HEADER_START_COL + j, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if j in (0, 1, 9) else LEFT_WRAP

    last_row = DATA_START_ROW + max(len(tcs), 1) - 1

    prio_range = f"{PRIORITY_COL}{DATA_START_ROW}:{PRIORITY_COL}{last_row}"
    ws.conditional_formatting.add(prio_range, CellIsRule(operator="equal", formula=['"high"'], fill=RED, font=WHITE_BOLD))
    ws.conditional_formatting.add(prio_range, CellIsRule(operator="equal", formula=['"mid"'], fill=YELLOW))
    ws.conditional_formatting.add(prio_range, CellIsRule(operator="equal", formula=['"low"'], fill=LIGHT_GREEN))

    if tcs:
        dv = DataValidation(type="list", formula1='"Pass,Fail,Block,N/A"', allow_blank=True, showDropDown=False)
        dv.add(f"{RESULT_COL}{DATA_START_ROW}:{RESULT_COL}{last_row}")
        ws.add_data_validation(dv)

    ws.row_dimensions[HEADER_ROW].height = 28
    for r in range(DATA_START_ROW, last_row + 1):
        ws.row_dimensions[r].height = 60


def add_summury_sheet(wb: Workbook, app_name: str, tab_data: list[tuple[str, int]]) -> None:
    ws = wb.create_sheet(title="summury", index=0)

    widths = {1: 3, 2: 22, 3: 12, 4: 8, 5: 8, 6: 8, 7: 8, 8: 9, 9: 9, 10: 9}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    title = ws.cell(row=2, column=2, value=f"{app_name} TC")
    title.font = Font(bold=True, size=14)

    # 환경 블록 (Row 4~7) — 빈 칸으로 제공, QA가 직접 기입
    env_rows = [("Android", "", "version", ""), ("iOS", "", "version", ""), ("OS", "", "version", "")]
    for j, label in enumerate(["수행 환경", "수행자", "테스트 기간", ""]):
        c = ws.cell(row=4, column=2 + j, value=label if label else None)
        if label:
            c.fill = PatternFill("solid", fgColor="00FF00")
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = BORDER
    for i, (env, who, _ver_label, ver) in enumerate(env_rows):
        r = 5 + i
        c1 = ws.cell(row=r, column=2, value=env); c1.fill = PatternFill("solid", fgColor="00FF00"); c1.font = HEADER_FONT; c1.alignment = CENTER; c1.border = BORDER
        c2 = ws.cell(row=r, column=3, value=who); c2.alignment = CENTER; c2.border = BORDER
        c3 = ws.cell(row=r, column=4, value="version"); c3.fill = PatternFill("solid", fgColor="00FF00"); c3.font = HEADER_FONT; c3.alignment = CENTER; c3.border = BORDER
        c4 = ws.cell(row=r, column=5, value=ver); c4.alignment = CENTER; c4.border = BORDER

    stat_header = ["구분", "검증 항목", "Pass", "Fail", "Block", "N/A", "성공율", "결함율", "수행율"]
    for j, label in enumerate(stat_header):
        c = ws.cell(row=9, column=2 + j, value=label)
        c.fill = GRAY; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER

    pct = "0.00%"
    fixed_rows = 9
    start = 10
    for i in range(fixed_rows):
        r = start + i
        for col in range(2, 11):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).alignment = CENTER
        if i < len(tab_data):
            tab, count = tab_data[i]
            esc = tab.replace("'", "''")
            last = DATA_START_ROW + count - 1
            ref_b = f"'{esc}'!B{DATA_START_ROW}:B{last}"
            ref_k = f"'{esc}'!{RESULT_COL}{DATA_START_ROW}:{RESULT_COL}{last}"
            ws.cell(row=r, column=2, value=tab)
            ws.cell(row=r, column=3, value=f"=COUNTA({ref_b})")
            ws.cell(row=r, column=4, value=f'=COUNTIF({ref_k},"Pass")')
            ws.cell(row=r, column=5, value=f'=COUNTIF({ref_k},"Fail")')
            ws.cell(row=r, column=6, value=f'=COUNTIF({ref_k},"Block")')
            ws.cell(row=r, column=7, value=f'=COUNTIF({ref_k},"N/A")')
            ws.cell(row=r, column=8,  value=f"=IFERROR(D{r}/C{r},0)").number_format = pct
            ws.cell(row=r, column=9,  value=f"=IFERROR(E{r}/C{r},0)").number_format = pct
            ws.cell(row=r, column=10, value=f"=IFERROR((D{r}+E{r}+F{r})/C{r},0)").number_format = pct

    total_r = start + fixed_rows
    last_data_r = start + fixed_rows - 1
    c = ws.cell(row=total_r, column=2, value="총 합계"); c.fill = GRAY; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    for col_idx, formula_col in [(3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")]:
        cc = ws.cell(row=total_r, column=col_idx, value=f"=SUM({formula_col}{start}:{formula_col}{last_data_r})")
        cc.fill = GRAY; cc.font = HEADER_FONT; cc.alignment = CENTER; cc.border = BORDER
    cc = ws.cell(row=total_r, column=8,  value=f"=IFERROR(D{total_r}/C{total_r},0)");                          cc.number_format = pct; cc.fill = GRAY; cc.font = HEADER_FONT; cc.alignment = CENTER; cc.border = BORDER
    cc = ws.cell(row=total_r, column=9,  value=f"=IFERROR(E{total_r}/C{total_r},0)");                          cc.number_format = pct; cc.fill = GRAY; cc.font = HEADER_FONT; cc.alignment = CENTER; cc.border = BORDER
    cc = ws.cell(row=total_r, column=10, value=f"=IFERROR((D{total_r}+E{total_r}+F{total_r})/C{total_r},0)"); cc.number_format = pct; cc.fill = GRAY; cc.font = HEADER_FONT; cc.alignment = CENTER; cc.border = BORDER


# ── 진입점 ───────────────────────────────────────────────────────────────────

def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="TC JSON → XLSX 워크북")
    parser.add_argument("app_name", nargs="?", default="app", help="앱 이름")
    parser.add_argument("--spec-version", default="", help="명세서 버전 (예: v1.2)")
    args = parser.parse_args(argv[1:])
    app_name = args.app_name

    today = date.today().strftime("%Y-%m-%d")
    deletion_note = f"명세 변경으로 삭제됨 ({today}{', ' + args.spec_version if args.spec_version else ''})"

    if not SRC_DIR.exists():
        print(f"[export_workbook] {SRC_DIR} 없음", file=sys.stderr)
        return 1

    sources = sorted(SRC_DIR.glob("*.json"))
    if not sources:
        print("[export_workbook] outputs/testcases/*.json 없음", file=sys.stderr)
        return 1

    all_tcs: list[dict] = []
    for s in sources:
        all_tcs.extend(load_json(s))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"{app_name}.xlsx"

    # [C1] merge_results → group_by_tab 순서 준수: result 병합 후 그룹화해야 시트에 반영됨
    snapshot = extract_result_snapshot(out_path)
    if snapshot:
        print(f"[export_workbook] 기존 result 스냅샷 로드: {len(snapshot)}개 TC", file=sys.stderr)
        all_tcs = merge_results(all_tcs, snapshot, deletion_note=deletion_note)

    groups = group_by_tab(all_tcs)
    if not groups:
        print("[export_workbook] 유효한 (screen, platform) 그룹 없음", file=sys.stderr)
        return 1

    wb = Workbook()
    wb.remove(wb.active)

    # [W1/W2] 정규화된 시트명을 tab_data에 저장해 summury 수식 참조와 일치시킴
    used_names: set[str] = set()
    tab_data: list[tuple[str, int]] = []
    for (screen, platform), tcs in sorted(groups.items()):
        sheet_name = normalize_sheet_name(f"{screen}_{platform}", used_names)
        used_names.add(sheet_name)
        add_tc_sheet(wb, sheet_name, tcs)
        tab_data.append((sheet_name, len(tcs)))

    add_summury_sheet(wb, app_name, tab_data)

    wb.save(out_path)
    save_snapshot(out_path, all_tcs)

    rel = out_path.relative_to(ROOT)
    total = sum(len(v) for v in groups.values())
    print(f"[export_workbook] wrote {rel}  ({len(tab_data)} sheets, {total} TC)")
    for name, count in tab_data:
        print(f"  - {name}  ({count} TC)")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
